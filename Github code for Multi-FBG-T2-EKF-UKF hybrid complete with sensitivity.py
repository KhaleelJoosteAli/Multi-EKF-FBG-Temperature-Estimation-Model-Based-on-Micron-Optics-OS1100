"""
This simulation framework demonstrates the integration of a realistic EV pouch cell model
with a bonded Micron Optics OS1100 FBG sensor for strain-compensated thermal measurement,
supported by three reference temperature FBGs. It incorporates thermal two-node (T2) dynamics,
simultaneous SoC and temperature estimation via EKF, UKF, and hybrid EKF-UKF approaches,
and realistic synthetic driving cycles spanning uphill, flat, and downhill road conditions.

The code allows for extensive calibration and parameter tuning, including process noise,
voltage measurement noise, FBG interrogator biases and drift, hysteresis corrections,
temperature- and current-dependent OCV/capacity adjustments, and slope-aware vehicle power mapping.
Monte Carlo analysis can be easily adjusted to explore estimator
robustness, sensitivity, and performance metrics (ME, MAE, RMSE) over a wide temperature range (-20 to 60).

Users can modify cycle names, time vectors, slope percentages, or FBG/thermal parameters
to explore alternate operating scenarios or vehicle configurations. This flexibility makes
the framework suitable for both research-oriented validation studies and practical
performance assessment of FBG-based SoC and thermal estimation strategies under realistic EV conditions.

The code’s modular structure and comprehensive commenting ensure that new synthetic cycles,
sensor configurations, or estimator variants can be incorporated with minimal changes,
allowing rapid iteration and exploration of future improvements, including AI integration
or experimental validation.
"""

# ----------------------------- IMPORTS + GLOBALS -----------------------------
import os
import math
import time
import json
import glob
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------- OUTPUT / PATHS ----------
output_folder = "simulation_outputs"
plots_folder = os.path.join(output_folder, "plots")
timeseries_folder = os.path.join(output_folder, "timeseries")
summary_folder = os.path.join(output_folder, "summary")
os.makedirs(plots_folder, exist_ok=True)
os.makedirs(timeseries_folder, exist_ok=True)
os.makedirs(summary_folder, exist_ok=True)
print("\n=== COLUMN DIAGNOSTIC ===")

# ---------- USER / SIM CONFIG ----------
seeded = False
seed_value = 10

# Monte Carlo configuration
num_mc_runs = 1   # reduce to e.g. 10 for quick tests
dt = 1.0          # s timestep
simulation_time = 2000.0   # seconds per run
time_vec = np.arange(0.0, simulation_time, dt)

# Ambient sweep (example); keep small for initial tests
T_amb_range = np.arange(25, 26, 1)   # e.g., [25] for development; expand to [-20, -10, 0, ... , 60] later

# ----------------------------- VEHICLE / BATTERY PROFILE (60 Ah pouch) -----------------------------
# This block defines the nominal electrical and vehicle-level parameters used to
# couple the lithium-ion pouch cell model to realistic EV driving conditions.
# The values are selected to represent a mid-size passenger EV and a single
# high-capacity pouch cell operating within a high-voltage battery pack.

# Nominal cell capacity
capacity_ah = 60.0
# Rated capacity of the lithium-ion pouch cell in ampere-hours.
# This value is representative of modern EV pouch cells and directly governs
# the coulomb-counting dynamics used in SoC estimation.

Q_n = capacity_ah * 3600.0
# Total usable charge in coulombs, derived from the nominal capacity.
# This quantity is used in the SoC state equation to integrate current over time
# and convert electrical load into SoC depletion or recovery.

# Nominal electrical characteristics
nominal_cell_voltage = 3.7
# Typical nominal voltage of a lithium-ion pouch cell.
# This value provides a realistic reference for cell-level behaviour but is not
# directly used for SoC estimation,

pack_nominal_voltage = 400.0
# Nominal high-voltage DC bus of the EV battery pack.
# This value is used to convert the mechanical power demand at the vehicle level
# into an equivalent electrical current drawn from or returned to the battery.
# A 400 V architecture is representative of mainstream EV platforms and ensures
# that simulated current magnitudes remain realistic across all drive cycles.
# Although the electro-thermal model operates at cell level, this parameter
# provides a physically consistent scaling between vehicle power and battery current.

# ----------------------------- VEHICLE / POWER MAPPING DEFAULTS -----------------------------

vehicle_mass_kg = 1500.0
# Total vehicle mass including chassis, battery pack, drivetrain, and nominal payload.
# Vehicle mass directly influences inertial forces during acceleration and deceleration,
# thereby shaping transient traction power demand and corresponding battery current.
# This parameter is essential for reproducing realistic current spikes under aggressive
# driving cycles such as US06 and DTS.

C_rr_default = 0.01
# Rolling resistance co-efficient for low rolling resistance passenger vehicle tyres.
# Rolling resistance introduces a continuous load proportional to vehicle weight and speed,
# representing mechanical losses due to tyre deformation and road interaction.
# Including this term ensures that steady-state cruising conditions still impose
# a non-zero electrical load on the battery.

rho_air = 1.225
# Air density under standard atmospheric conditions at sea level (kg/m^3).
# This parameter is required for aerodynamic drag calculations and becomes dominant
# at medium to high vehicle speeds. Accurate representation of air density ensures
# realistic power demand during highway and high-speed drive cycles.

A_f_default = 2.2
# Effective frontal area of the vehicle (m^2).
# The frontal area determines the magnitude of aerodynamic drag acting on the vehicle
# and directly affects power demand as vehicle speed increases.
# This value reflects a typical sedan or compact SUV configuration.

Cd_default = 0.32
# Aerodynamic drag co-efficient of the vehicle.
# The drag co-efficient captures the efficiency of the vehicle body shape with respect
# to airflow and, together with frontal area and air density, defines aerodynamic losses.
# A value of 0.32 is consistent with modern production EVs designed for efficiency.
# This parameter is particularly influential during sustained high-speed operation.


# ----------------------------- FBG / SENSOR / MODEL GLOBALS -----------------------------
# Number of FBG channels and which index is bonded
num_fbgs = 4
bonded_index = 0   # index of the FBG that is bonded to the cell surface (0..num_fbgs-1)

# FBG silica & thermo-optic coefficients 
lambda_B0 = 1550e-9        # m
rho_e = 0.6                # effective photoelastic constant 
alpha_fbg = 0.55e-6        # coefficient of thermal expansion of fibre (1/C) 
xi = 6.7e-6                # thermo-optic coefficient (1/C)

# Convenience
pm_to_m = 1e-12

# ----------------------------- ELECTRICAL / THERMAL NOMINALS -----------------------------
# Base resistances for the Thevenin equivalent circuit (Ohms)
R0_ref_base = 0.015       # Series resistance (cell bulk)
R1_ref_base = 0.01        # First polarization resistance
R2_ref_base = 0.02        # Second polarization resistance

# Thevenin RC capacitances (Farads)
C1_base = 2400.0          # First RC branch capacitance
C2_base = 4800.0          # Second RC branch capacitance

# Reference temperature for all nominal calculations (degrees Celsius)
T_ref = 25.0

# Thermal two-node (T2) model nominal values for a 60 Ah pouch cell
C_th_core_nominal = 400.0    # Thermal capacity of the cell core
C_th_surf_nominal = 200.0    # Thermal capacity of the cell surface
R_th_nominal = 3.0            # Overall thermal resistance to ambient
R_th_core_surf = 0.8          # Thermal resistance between core and surface (coarse approximation)

# ----------------------------- REAL EV ADJUSTMENTS (AGREED) -----------------------------
# Temperature coefficient of resistance, capturing SoC and ageing effects
alpha_R_base = 0.0085         # Nominal temperature coefficient (1/C), within typical range [0.006 .. 0.018]
alpha_R_min = 0.006           # Minimum expected coefficient
alpha_R_max = 0.018           # Maximum expected coefficient
alpha_R_soc = -0.15           # SoC-dependent adjustment
alpha_R_aging = 0.35          # Ageing multiplier for long-term drift

# FBG bonding and strain transfer parameters
k_th_nominal = 8e-7           # Mechanical strain coefficient per degree Celsius (reduced from 15e-6 for realism)
eta_bond_nominal = 0.8        # Nominal bond transfer efficiency (realistic EV range 0.6 .. 0.9)
eta_bond_min = 0.6            # Minimum feasible bond transfer
eta_bond_max = 1.2            # Maximum feasible bond transfer
tau_creep = 48.0 * 3600.0     # Creep time constant in seconds (48 hours)
eta_bond_creep_drop = 0.01    # Long-term decrease in bond transfer

# Thermal parameter uncertainties applied per Monte Carlo run
sigma_C_frac = 0.03           # Fractional uncertainty in thermal capacitance
sigma_Rth_frac = 0.05         # Fractional uncertainty in thermal resistance

# ----------------------------- INTERROGATOR / FBG NOISE / DRIFT -----------------------------
# FBG measurement parameters and uncertainties
# All wavelength-related units are initially in picometers (pm) for convenience; conversion to meters applied when used

fbg_bias_pm_mu = 0.0                 # Mean static bias of FBG interrogator, pm
fbg_bias_pm_sigma = 2.0              # Standard deviation of static bias uncertainty, pm
fbg_drift_pm_per_hour = 0.1          # Slow temporal drift of the FBG interrogator, pm/hour
fbg_adc_noise_pm = 3.0               # High-frequency per-sample ADC noise, pm
fbg_multiplex_jitter_ms = 2.0        # Timing jitter in multiplexed interrogation, milliseconds
fbg_quantization_pm = 0.1            # ADC quantization resolution, pm; set to 0 to disable quantization

# Cross-talk between surface and core FBG channels
c_TS_nominal = 0.15                  # Fraction of signal leakage; realistic range 0.05..0.15

# ----------------------------- PROCESS / MEASUREMENT NOISE CONFIG -----------------------------
# Process noise covariance matrix for EKF/UKF
# Slightly inflated to account for model uncertainties and unmodelled dynamics
Q_process = np.diag([3.0e-5, 2e-5, 2e-5, 3e-3, 3e-3])

# Voltage sensor measurement noise variance (squared)
R_voltage = (3.0e-3)**2

# High-frequency FBG noise converted to meters for simulation
R_fbg_wavelength = (fbg_adc_noise_pm * pm_to_m)**2

# ----------------------------- UKF / EKF PARAMETERS -----------------------------
# Tuning parameters for Unscented Kalman Filter (UKF)
UKF_ALPHA = 1e-3                     # Spread of sigma points around mean
UKF_BETA = 2.0                       # Optimal for Gaussian distributions
UKF_KAPPA = 0.0                       # Secondary scaling parameter


COLORS = {"EKF": "tab:blue", "UKF": "tab:orange", "EKF-UKF-hybrid": "tab:green"}

# ----------------------------- REFINEMENT PARAMETERS (ADDITIVE) -----------------------------
# Temperature dependent capacity / OCV corrections
# capacity_temp_coef is fractional capacity change per degree C (small)
capacity_temp_coef = -2.2e-4    # -0.0005 per degC (−0.05% per C, small realistic effect)
# OCV temperature coefficient (V per degC). Typical range around -0.0005 to -0.001 V/C
ocv_temp_coef = -2.0e-4         # -0.0006 V per degC

# Current dependent polarization resistance correction
# k_I_polar reduces effective polarization resistance at higher currents.
# Factor applied as multiplier: R_eff = R_base * (1 + ...) / (1 + k_I_polar * |I|)
k_I_polar = 7e-3              # 0.005 per amp. (at 100 A gives reduction factor ~1/(1+0.5) = 0.667)

# Adaptive measurement noise in low SoC region
# low_soc_threshold defines below which SoC fraction the noise scaling begins.
low_soc_threshold = 0.15      # below 20% SoC noise starts to increase
low_soc_max_scale = 2.0       # measurement noise can increase up to 5x at SoC = 0

# Hysteresis correction simplified model (forward-backward relaxation)
# h0 is maximum steady-state hysteresis magnitude in volts (approx 5-30 mV depending on chemistry)
hyst_h0 = 0.015              # 20 mV nominal hysteresis magnitude
hyst_kI = 0.008              # A^-1 shape factor for tanh(I) (so hysteresis grows with |I|)
hyst_tau = 60.0               # seconds relaxation time constant for forward/backward behaviour

def vehicle_power_from_speed(speed_ms,
                             mass_kg=vehicle_mass_kg,
                             C_rr=C_rr_default,
                             rho_air_local=rho_air,
                             A_f=A_f_default,
                             Cd=Cd_default,
                             slope_percent= 0.0):
    """
    Computes the instantaneous vehicle power demand from speed (m/s), including:
    - Rolling resistance (F_roll)
    - Aerodynamic drag (F_aero)
    - Acceleration (F_acc)
    - Road gradient (slope_percent, positive for uphill, negative for downhill)

    The slope is expressed as a percentage; small-angle approximation is applied.
    Acceleration is limited to [-5, +5] m/s^2 to avoid unrealistically high forces.
    Returns vehicle power in Watts.
    """
    # Compute numerical acceleration from speed profile
    acc = np.gradient(speed_ms, edge_order=2) if isinstance(speed_ms, np.ndarray) else 0.0
    acc = np.clip(acc, -5.0, 5.0)

    # Rolling resistance force: F = C_rr * m * g
    F_roll = C_rr * mass_kg * 9.81

    # Aerodynamic drag: F = 0.5 * rho * Cd * A * v^2
    F_aero = 0.5 * rho_air_local * Cd * A_f * speed_ms**2

    # Acceleration-induced force: F = m * a
    F_acc = mass_kg * acc

    # Slope-induced force: F = m * g * sin(theta) ≈ m * g * slope_percent / 100
    g = 9.81
    F_slope = mass_kg * g * (slope_percent / 100.0)

    # Total longitudinal force
    F_total = F_roll + F_aero + F_acc + F_slope

    # Instantaneous vehicle power: P = F_total * velocity
    P_vehicle = F_total * speed_ms
    return P_vehicle


def profile_to_current(speed_ms, V_pack=pack_nominal_voltage, eta_drive=0.9, regen_eta=0.85):
    """
    Converts a vehicle speed profile to battery current (A) for the EV model.
    Accounts for:
    - Drive efficiency (eta_drive)
    - Regenerative braking efficiency (regen_eta)
    - Maximum allowed pack power (10x nominal capacity voltage*Ah)

    Negative power (braking) is scaled by regen_eta; positive power is scaled by eta_drive.
    Returns array of battery currents corresponding to speed profile.
    """
    P_vehicle = vehicle_power_from_speed(speed_ms)

    # Clip power to realistic max limits to prevent unrealistic currents
    max_power = V_pack * capacity_ah * 10.0
    P_vehicle = np.clip(P_vehicle, -max_power, max_power)

    # Convert vehicle power to battery current
    I = P_vehicle / V_pack / eta_drive

    # Apply regenerative braking efficiency to negative currents
    if isinstance(P_vehicle, np.ndarray):
        I[P_vehicle < 0] = P_vehicle[P_vehicle < 0] * regen_eta / V_pack
    else:
        if P_vehicle < 0:
            I = P_vehicle * regen_eta / V_pack

    return I
# ----------------------------- SYNTHETIC DRIVING CYCLES WITH SLOPE -----------------------------
def generate_synthetic_profile(name, t_vec, slope_percent=0.0):
    """
    Generates current and speed profiles for synthetic EV driving cycles. 
    Each cycle is represented as a sinusoidal or quasi-realistic speed profile to emulate real-world driving conditions.

    Parameters
    ----------
    name : str
        Name of the driving cycle (e.g., "UDDS", "FUDS", "FTP10Hz", "HWFET", "US06", "DTS", "HIGH_SPEED").
    t_vec : array_like
        Time vector in seconds over which to generate the profile.
    slope_percent : float, optional
        Road gradient as a percentage. Positive for uphill, negative for downhill.

    Returns
    -------
    current : ndarray
        Battery current profile (A) corresponding to the speed profile.
    speed_ms : ndarray
        Speed profile in meters per second.

    Notes on cycles:
    - UDDS: Urban Dynamometer Driving Schedule, representing low-speed, stop-and-go city traffic with gentle acceleration.
    - FUDS: Federal Urban Driving Schedule, similar to UDDS but with slightly higher average speed and smoother sinusoidal variations.
    - FTP10Hz: Fast Transient Profile, high-frequency urban acceleration/deceleration events, capturing short-term dynamic loads on the battery.
    - HWFET: Highway Fuel Economy Test, representing steady high-speed highway driving with moderate sinusoidal fluctuations to emulate small speed changes.
    - HWFET10Hz: High-frequency variant of HWFET, capturing faster speed oscillations common in congested highways or dynamic overtaking scenarios.
    - US06: Aggressive US06 cycle, representing rapid acceleration and deceleration events on urban/highway mixed driving.
    - DTS: Dynamic test scenario with added "spikes" to simulate hard accelerations or braking events, mimicking performance-oriented driving.
    - HIGH_SPEED: Constant high-speed highway profile with small sinusoidal variations, representing sustained high-power cruising.

    The generated speed is clamped to be non-negative and converted to m/s. 
    Battery current is then mapped from speed using vehicle power calculations, including slope effects and drive/regenerative efficiencies.
    """

    # Define sinusoidal speed profiles for each cycle
    if name == "UDDS":
        speed_kmh = 36.0 + 12.0 * np.sin(2 * np.pi * t_vec / 600.0)
    elif name == "FUDS":
        speed_kmh = 40.0 + 15.0 * np.sin(2 * np.pi * t_vec / 500.0)
    elif name == "FTP10Hz":
        speed_kmh = 30.0 + 10.0 * np.sin(2 * np.pi * t_vec / 200.0)
    elif name == "HWFET":
        speed_kmh = 72.0 + 25.0 * np.sin(2 * np.pi * t_vec / 800.0)
    elif name == "HWFET10Hz":
        speed_kmh = 65.0 + 18.0 * np.sin(2 * np.pi * t_vec / 150.0)
    elif name == "US06":
        speed_kmh = 54.0 + 30.0 * np.sin(2 * np.pi * t_vec / 400.0)
    elif name == "DTS":
        speed_kmh = 62.0 + 28.0 * np.sin(2 * np.pi * t_vec / 350.0)
        # Add occasional high acceleration spikes to emulate aggressive driving
        spikes = (np.sin(2 * np.pi * t_vec / 60.0) > 0.98).astype(float) * 60.0
        speed_kmh += spikes
    elif name == "HIGH_SPEED":
        speed_kmh = 110.0 + 10.0 * np.sin(2 * np.pi * t_vec / 900.0)
    else:
        speed_kmh = np.zeros_like(t_vec)

    # Clamp negative speeds to zero (no reverse motion)
    speed_kmh = np.maximum(0.0, speed_kmh)

    # Convert speed from km/h to m/s
    speed_ms = speed_kmh / 3.6

    # Nested function to convert speed profile to battery current, including slope
    def profile_to_current_with_slope(speed_ms, V_pack=pack_nominal_voltage, eta_drive=0.9, regen_eta=0.85):
        """
        Converts speed to battery current considering:
        - Vehicle power from speed and slope
        - Maximum allowable pack power
        - Drive and regenerative braking efficiency
        """
        P_vehicle = vehicle_power_from_speed(speed_ms, slope_percent=slope_percent)
        max_power = V_pack * capacity_ah * 10.0
        P_vehicle = np.clip(P_vehicle, -max_power, max_power)
        I = P_vehicle / V_pack / eta_drive

        # Apply regenerative braking efficiency for negative power
        if isinstance(P_vehicle, np.ndarray):
            I[P_vehicle < 0] = P_vehicle[P_vehicle < 0] * regen_eta / V_pack
        else:
            if P_vehicle < 0:
                I = P_vehicle * regen_eta / V_pack
        return I

    # Generate corresponding current profile
    current = profile_to_current_with_slope(speed_ms)

    return current, speed_ms

# Updated cycle names
cycle_names = ["UDDS", "FUDS", "FTP10Hz", "HWFET", "HWFET10Hz", "US06", "DTS", "HIGH_SPEED"]

# ----------------------------- OCV LUT -----------------------------
def ocv_soc_equation(SoC_frac):
    return 3.0 + 1.2 * SoC_frac - 0.1 * (SoC_frac**2)

_soc_percent_lut = np.arange(0, 101, 1)
_soc_frac_lut = _soc_percent_lut / 100.0
_ocv_lut_values = ocv_soc_equation(_soc_frac_lut)

def ocv_from_lut(SoC_frac):
    SoC_clamped = np.clip(SoC_frac, 0.0, 1.0)
    return np.interp(SoC_clamped, _soc_frac_lut, _ocv_lut_values)

# ----------------------------- TEMPERATURE DEPENDENT OCV / CAPACITY -----------------------------
def capacity_temp_factor(T_cell):
    """
    Returns multiplicative factor for nominal capacity Q_n based on cell temperature.
    Uses simple linear approximation around T_ref. Values chosen as small realistic effects.
    """
    return 1.0 + capacity_temp_coef * (T_cell - T_ref)

def ocv_temp_correction(ocv, T_cell):
    """
    Apply a linear temperature correction to OCV.
    ocv: baseline open circuit voltage from LUT
    T_cell: cell/core temperature in degC used for correction
    Returns corrected OCV in volts.
    """
    return ocv + ocv_temp_coef * (T_cell - T_ref)

def ocv_from_lut_temp(SoC_frac, T_cell):
    """
    Wrapper that returns OCV corrected for temperature.
    Keeps original LUT based behaviour and applies a small temperature-dependent offset.
    """
    base = ocv_from_lut(SoC_frac)
    return ocv_temp_correction(base, T_cell)

# ============================================================
# PLOT OCV–SOC CURVES WITH TEMPERATURE CORRECTION AND R²
# ============================================================

def plot_ocv_soc_lut_with_r2_manual():
    # SoC range for smooth plotting
    soc_plot = np.linspace(0, 1, 200)

    # Baseline OCV (no temperature effect)
    ocv_baseline = ocv_from_lut(soc_plot)

    # Temperatures from -20 to 60 in 10-degrees increments
    temps = np.arange(-20, 61, 10)  # [-20, -10, 0, 10, ..., 60]

    plt.figure(figsize=(8, 5))

    # Plot baseline LUT
    plt.plot(soc_plot * 100, ocv_baseline,
             label="Baseline OCV (T = T_ref)",
             linewidth=2,
             color='black')

    # Store R² annotations
    r2_text_lines = []

    # Plot temperature-adjusted curves for each temperature and compute R²
    for T in temps:
        ocv_temp = ocv_from_lut_temp(soc_plot, T)
        plt.plot(soc_plot * 100, ocv_temp,
                 label=f"OCV with T = {T} °C")
        
        # Compute R² manually
        y_mean = np.mean(ocv_temp)
        ss_tot = np.sum((ocv_temp - y_mean) ** 2)
        ss_res = np.sum((ocv_temp - ocv_baseline) ** 2)
        r2 = 1 - ss_res / ss_tot
        r2_text_lines.append(f"T={T:>3}°C: R²={r2:.4f}")

    plt.xlabel("State of Charge (%)")
    plt.ylabel("Open Circuit Voltage (V)")
    plt.title("OCV–SOC Look-Up Table with Temperature Dependence")
    plt.grid(True)
    plt.legend()

    # Add R² annotations in right-hand bottom with proper spacing
    x_text = 80  # percent SoC
    y_text_start = min(ocv_baseline) + 0.02  # slightly above min OCV
    line_height = 0.05  # vertical spacing between lines

    for i, line in enumerate(r2_text_lines):
        plt.text(x_text, y_text_start + i * line_height, line, fontsize=9, ha='left', va='bottom')

    plt.tight_layout()
    plt.show()

# ============================================================
# PRINT LUT TABLE TO CONSOLE
# ============================================================

# Define SoC points for printing
soc_points = np.linspace(0, 1, 11)  # 0%, 10%, ..., 100%

print("State of Charge (%) | Baseline OCV (V) | Temperature-corrected OCV (25°C)")
print("-" * 60)
for soc in soc_points:
    soc_percent = soc * 100
    ocv_base = ocv_from_lut(soc)
    ocv_temp25 = ocv_from_lut_temp(soc, 20.0)  # example at 25°C
    print(f"{soc_percent:>6.1f}             | {ocv_base:>6.3f}         | {ocv_temp25:>6.3f}")

# Print message to console
print("Generating OCV–SoC curves for baseline and temperature-corrected OCVs with R²...")

# Call the plotting function
plot_ocv_soc_lut_with_r2_manual()

# Print completion message
print("OCV–SoC curves displayed successfully.")

# ============================================================
# CALL PLOTTING FUNCTION AND PRINT LUT DATA
# ============================================================

# Define SoC points for printing
soc_points = np.linspace(0, 1, 11)  # 0%, 10%, ..., 100%

print("State of Charge (%) | Baseline OCV (V) | Temperature-corrected OCV (25°C)")
print("-" * 60)
for soc in soc_points:
    soc_percent = soc * 100
    ocv_base = ocv_from_lut(soc)
    ocv_temp25 = ocv_from_lut_temp(soc, 20.0)  # example at 25°C
    print(f"{soc_percent:>6.1f}             | {ocv_base:>6.3f}         | {ocv_temp25:>6.3f}")

# Print message to console
print("Generating OCV–SoC curves for baseline and temperature-corrected OCVs...")


# Print completion message
print("OCV–SoC curves displayed successfully.")

# ----------------------------- UTILS: PSD, CHOLESKY, SIGMA POINTS -----------------------------
def make_psd(P, min_eig=1e-12):
    P = 0.5 * (P + P.T)
    vals, vecs = np.linalg.eigh(P)
    vals_clipped = np.clip(vals, min_eig, None)
    P_psd = (vecs @ np.diag(vals_clipped)) @ vecs.T
    return 0.5 * (P_psd + P_psd.T)

def safe_cholesky(M, initial_eps=1e-12, max_tries=12):
    M = 0.5 * (M + M.T)
    eps = initial_eps
    for _ in range(max_tries):
        try:
            return np.linalg.cholesky(M + np.eye(M.shape[0]) * eps)
        except np.linalg.LinAlgError:
            eps *= 10.0
    vals, vecs = np.linalg.eigh(M)
    vals_clipped = np.clip(vals, 1e-12, None)
    M_fixed = (vecs @ np.diag(vals_clipped)) @ vecs.T
    return np.linalg.cholesky(0.5 * (M_fixed + M_fixed.T))

def generate_sigma_points(x, P, alpha=UKF_ALPHA, beta=UKF_BETA, kappa=UKF_KAPPA, min_c=1e-6):
    n = len(x)
    lam = alpha**2 * (n + kappa) - n
    c = n + lam
    c = max(c, min_c)
    P_safe = make_psd(P)
    sqrt_mat = safe_cholesky(P_safe * c)
    sqrtP_cols = sqrt_mat.T
    sigma = np.zeros((2 * n + 1, n))
    sigma[0] = x.copy()
    for i in range(n):
        sigma[i + 1] = x + sqrtP_cols[:, i]
        sigma[i + 1 + n] = x - sqrtP_cols[:, i]
    Wm = np.full(2 * n + 1, 1.0 / (2 * c))
    Wc = np.full(2 * n + 1, 1.0 / (2 * c))
    Wm[0] = lam / c
    Wc[0] = lam / c + (1.0 - alpha**2 + beta)
    return sigma, Wm, Wc

def compute_jacobian(f, x, I, T_amb_local, eps=1e-6):
    n = len(x)
    f0 = f(x, I, T_amb_local)
    J = np.zeros((n, n))
    for i in range(n):
        dx = np.zeros_like(x)
        dx[i] = eps
        J[:, i] = (f(x + dx, I, T_amb_local) - f0) / eps
    return J

def compute_measurement_jacobian(h, x, I, eps=1e-6):
    h0 = h(x, I)
    m = len(h0)
    n = len(x)
    H = np.zeros((m, n))
    for i in range(n):
        dx = np.zeros_like(x)
        dx[i] = eps
        H[:, i] = (h(x + dx, I) - h0) / eps
    return H

# ----------------------------- SECTION: NOISE & PARAM SAMPLING -----------------------------
def sample_log_normal_factor(mean=1.0, sigma_frac=0.05):
    return math.exp(np.random.normal(0.0, sigma_frac))

def sample_thermal_parameters():
    C_core_eff = C_th_core_nominal * sample_log_normal_factor(sigma_frac=sigma_C_frac)
    C_surf_eff = C_th_surf_nominal * sample_log_normal_factor(sigma_frac=sigma_C_frac)
    R_th_eff = R_th_nominal * sample_log_normal_factor(sigma_frac=sigma_Rth_frac)
    return float(C_core_eff), float(C_surf_eff), float(R_th_eff)

def sample_ageing_factor(max_age=0.25):
    return float(np.random.uniform(0.0, max_age))

def sample_fbg_interrogator_characteristics():
    b_lambda_pm = np.random.normal(fbg_bias_pm_mu, fbg_bias_pm_sigma)
    b_lambda_m = b_lambda_pm * pm_to_m
    eta_bond_0 = float(np.random.uniform(eta_bond_min, eta_bond_max))
    c_TS = float(np.random.uniform(0.05, 0.15))
    b0_pm = np.random.normal(0.0, 1.0)
    b0_m = b0_pm * pm_to_m
    drift_rate_pm_per_s = fbg_drift_pm_per_hour / 3600.0
    drift_rate_m_per_s = drift_rate_pm_per_s * pm_to_m
    quant_m = fbg_quantization_pm * pm_to_m if fbg_quantization_pm > 0 else 0.0
    return {
        "b_lambda_m": b_lambda_m,
        "eta_bond_0": eta_bond_0,
        "c_TS": c_TS,
        "b0_m": b0_m,
        "drift_rate_m_per_s": drift_rate_m_per_s,
        "quant_m": quant_m
    }

def apply_quantization(value, quant):
    if quant <= 0:
        return value
    return np.round(value / quant) * quant

# ----------------------------- SECTION: FBG MODEL (realistic) -----------------------------
def eta_bond_time(eta0, t_seconds, tau_creep_local=tau_creep, creep_drop=eta_bond_creep_drop):
    return float(max(0.0, eta0 - creep_drop * (1.0 - math.exp(-t_seconds / tau_creep_local))))

def fbg_wavelength_shift_realistic(T, eps=0.0, deltaT=None, c_TS=None):
    delta_lambda_base = lambda_B0 * ((1 - rho_e) * eps + (alpha_fbg + xi) * (T - T_ref))
    cross_term = 0.0
    if (c_TS is not None) and (deltaT is not None):
        cross_term = lambda_B0 * c_TS * deltaT * eps
    return delta_lambda_base + cross_term

# ----------------------------- RESISTANCE: TEMP + SOC + AGE + CURRENT DEPENDENCE -----------------------------
def resistance_T_realistic(R_ref, T, SoC_frac=0.5, ageing_factor=0.1, alpha_T=alpha_R_base,
                           alpha_soc=alpha_R_soc, alpha_age=alpha_R_aging, I=0.0):
    """
    Extended resistance model:
      R = R_ref * (1 + alpha_T*(T - T_ref) + alpha_soc*(SoC - 0.5) + alpha_age*ageing)
      then apply a current-dependent polarization reduction:
      R_eff = R / (1 + k_I_polar * abs(I))
    I is optional (default 0), so older calls without I remain valid.
    """
    R = R_ref * (1.0 + alpha_T * (T - T_ref) + alpha_soc * (SoC_frac - 0.5) + alpha_age * ageing_factor)
    # ensure R positive and apply current dependent scaling
    R = max(R, 1e-6)
    current_scaling = 1.0 / (1.0 + k_I_polar * abs(I))
    return float(R * current_scaling)

# ----------------------------- MEASUREMENT MODEL (UPDATED: OCV temp + hysteresis + current) -----------------------------
def measurement_model_realistic(x, I, t_seconds, interrogator_params, k_eff=None):
    """
    Returns measured vector [Vt, delta_lambda_1, ..., delta_lambda_N] and true delta_lambdas_true
    Changes:
      - uses ocv_from_lut_temp(SoC, Tc)
      - adds a simple forward/backward relaxation hysteresis term to the OCV used in the model
      - uses resistance_T_realistic with current dependence where appropriate
    """
    SoC, V1, V2, Tc, Ts = x

    # Effective resistances (current dependent)
    R0 = resistance_T_realistic(R0_ref_base, Tc, SoC_frac=SoC, ageing_factor=0.0, alpha_T=alpha_R_base, I=I)

    # OCV corrected by cell temperature
    ocv = ocv_from_lut_temp(SoC, Tc)

    # Simplified hysteresis term (same formula used for true and predicted measurements)
    # forward/backward relaxation approximated as steady-state scaled by (1 - exp(-t/tau))
    # sign preserved via tanh and sign(I)
    if abs(I) < 1e-6:
        # small currents produce negligible steady-state hysteresis
        h_ss = 0.0
    else:
        h_ss = hyst_h0 * np.tanh(hyst_kI * I)  # signed steady state hysteresis
    h_term = float(h_ss * (1.0 - math.exp(-t_seconds / hyst_tau)))

    # Terminal voltage with hysteresis added to OCV
    Vt = ocv + h_term - I * R0 - V1 - V2

    b_lambda_m = interrogator_params["b_lambda_m"]
    b0_m = interrogator_params["b0_m"]
    drift_rate_m_per_s = interrogator_params["drift_rate_m_per_s"]
    quant_m = interrogator_params["quant_m"]
    c_TS = interrogator_params["c_TS"]
    eta_bond_0 = interrogator_params["eta_bond_0"]

    # bonding evolution
    eta_bond_t = eta_bond_time(eta_bond_0, t_seconds)
    if k_eff is None:
        k_eff = k_th_nominal * eta_bond_t
    eps_bonded = k_eff * (Ts - T_ref)

    delta_lambdas_true = np.zeros(num_fbgs)
    delta_lambdas_meas = np.zeros(num_fbgs)

    # iterate FBGs (unchanged except for usage of Ts for deltaT)
    for i in range(num_fbgs):
        eps = eps_bonded if i == bonded_index else 0.0
        deltaT = Ts - T_ref
        delta_l_true = fbg_wavelength_shift_realistic(Ts, eps=eps, deltaT=deltaT, c_TS=c_TS)
        delta_lambdas_true[i] = delta_l_true

        drift_m = b0_m + drift_rate_m_per_s * t_seconds
        hf_noise_m = np.random.normal(0.0, fbg_adc_noise_pm * pm_to_m)
        measured = delta_l_true + b_lambda_m + drift_m + hf_noise_m
        measured = apply_quantization(measured, quant_m)

        if i != bonded_index:
            nonbond_jitter_pm = np.random.normal(0.0, 2.0)
            measured += nonbond_jitter_pm * pm_to_m

        delta_lambdas_meas[i] = measured

    return np.concatenate(([Vt], delta_lambdas_meas)), np.concatenate(([Vt], delta_lambdas_true))

# ----------------------------- STATE TRANSITION (UPDATED: capacity temp factor + I->resistance) -----------------------------
def state_transition_realistic(x, I, T_amb_local, C_core_eff, C_surf_eff, R_th_eff, ageing_factor):
    SoC, V1, V2, Tc, Ts = x

    # Effective resistances include explicit current dependence
    R0 = resistance_T_realistic(R0_ref_base, Tc, SoC_frac=SoC, ageing_factor=ageing_factor, alpha_T=alpha_R_base, I=I)
    R1 = resistance_T_realistic(R1_ref_base, Tc, SoC_frac=SoC, ageing_factor=ageing_factor, alpha_T=alpha_R_base, I=I)
    R2 = resistance_T_realistic(R2_ref_base, Tc, SoC_frac=SoC, ageing_factor=ageing_factor, alpha_T=alpha_R_base, I=I)

    # Thermal dynamics (unchanged)
    dTc = (I**2 * R0 - (Tc - Ts) / R_th_core_surf) / C_core_eff
    Tc_next = np.clip(Tc + dTc * dt, -50.0, 100.0)

    dTs = (Tc - Ts) / R_th_core_surf / C_surf_eff + (T_amb_local - Ts) / (R_th_eff * C_surf_eff)
    Ts_next = np.clip(Ts + dTs * dt, -50.0, 100.0)

    # Thevenin RC dynamics (unchanged structure)
    dV1 = (-V1 / (R1 * C1_base) + I / C1_base)
    dV2 = (-V2 / (R2 * C2_base) + I / C2_base)

    # Effective capacity corrected by temperature and ageing
    Q_eff = Q_n * capacity_temp_factor(Tc) * (1.0 - float(np.clip(ageing_factor, 0.0, 0.95)))
    Q_eff = max(Q_eff, Q_n * 0.5)  # enforce a sensible lower bound (50% of nominal) to avoid numeric issues

    # SoC update uses Q_eff
    dSoC = -I * dt / Q_eff if I >= 0 else -I * 0.995 * dt / Q_eff

    return np.array([np.clip(SoC + dSoC, 0.0, 1.0), V1 + dV1 * dt, V2 + dV2 * dt, Tc_next, Ts_next])

# ----------------------------- MEASUREMENT COVARIANCE (UPDATED: adaptive low-SoC noise) -----------------------------
def build_R_measure(num_fbgs_local, fbg_adc_noise_pm_local, interrogator_params, include_bias_uncertainty=True, SoC_frac=None):
    """
    Builds measurement covariance matrix for voltage + N FBG channels.
    If SoC_frac provided and below low_soc_threshold, scale FBG variance up to low_soc_max_scale.
    """
    sigma_fbg_m = fbg_adc_noise_pm_local * pm_to_m

    # Scale factor for low SoC region
    scale = 1.0
    if SoC_frac is not None:
        if SoC_frac < low_soc_threshold:
            frac = (low_soc_threshold - SoC_frac) / max(1e-6, low_soc_threshold)
            scale = 1.0 + frac * (low_soc_max_scale - 1.0)  # scales from 1..low_soc_max_scale

    R_diag = [R_voltage]
    for _ in range(num_fbgs_local):
        var = (sigma_fbg_m * scale)**2
        if include_bias_uncertainty:
            b_var = (fbg_bias_pm_sigma * pm_to_m)**2
            drift_uncertainty = (interrogator_params["drift_rate_m_per_s"] * (simulation_time) / 6.0)**2
            var += b_var + drift_uncertainty
        R_diag.append(var)
    return np.diag(R_diag)

# ----------------------------- SECTION: EKF (updated signature) -----------------------------
def ekf_predict_update(x_prev, P_prev, z, I, T_amb, C_core_eff, C_surf_eff, R_th_eff, ageing_factor,
                       interrogator_params, current_time_seconds, R_measure=None, Qp=None):
    if Qp is None:
        Qp = Q_process
    P_prev = make_psd(P_prev)

    def f_wrapper(x_local, I_local, T_local):
        return state_transition_realistic(x_local, I_local, T_local, C_core_eff, C_surf_eff, R_th_eff, ageing_factor)

    x_pred = f_wrapper(x_prev, I, T_amb)
    F = compute_jacobian(f_wrapper, x_prev, I, T_amb)
    P_pred = F @ P_prev @ F.T + Qp

    def h_wrapper(x_local, I_local):
        meas, _ = measurement_model_realistic(x_local, I_local, current_time_seconds, interrogator_params, k_eff=None)
        return meas

    H = compute_measurement_jacobian(h_wrapper, x_pred, I)

    # Build adaptive measurement noise using predicted SoC if R_measure not supplied
    if R_measure is None:
        SoC_pred = np.clip(x_pred[0], 0.0, 1.0)
        R_measure = build_R_measure(num_fbgs, fbg_adc_noise_pm, interrogator_params, SoC_frac=SoC_pred)

    S = H @ P_pred @ H.T + R_measure
    try:
        K = P_pred @ H.T @ np.linalg.inv(S)
    except np.linalg.LinAlgError:
        K = P_pred @ H.T @ np.linalg.pinv(S)

    z_pred, _ = measurement_model_realistic(x_pred, I, current_time_seconds, interrogator_params, k_eff=None)
    x_upd = x_pred + K @ (z - z_pred)
    P_upd = (np.eye(len(x_prev)) - K @ H) @ P_pred
    return np.nan_to_num(x_upd), make_psd(P_upd)

# ----------------------------- SECTION: UKF (updated) -----------------------------
def ukf_predict_update(x_prev, P_prev, z, I, T_amb, C_core_eff, C_surf_eff, R_th_eff, ageing_factor,
                       alpha=UKF_ALPHA, beta=UKF_BETA, kappa=UKF_KAPPA,
                       R_measure=None, Qp=None, interrogator_params=None, current_time_s=0.0):
    x_prev = np.asarray(x_prev)
    P_prev = np.asarray(P_prev)
    if Qp is None:
        Qp = Q_process

    P_prev = make_psd(P_prev)
    z = np.asarray(z)
    sigma, Wm, Wc = generate_sigma_points(x_prev, P_prev, alpha, beta, kappa)

    X_pred = np.array([state_transition_realistic(s, I, T_amb, C_core_eff, C_surf_eff, R_th_eff, ageing_factor) for s in sigma])
    x_pred = np.sum(Wm[:, None] * X_pred, axis=0)

    # Build adaptive measurement noise using predicted SoC if R_measure not supplied
    if R_measure is None:
        SoC_pred = np.clip(x_pred[0], 0.0, 1.0)
        R_measure = build_R_measure(num_fbgs, fbg_adc_noise_pm, interrogator_params, SoC_frac=SoC_pred)

    n = x_prev.shape[0]
    P_pred = np.zeros((n, n))
    for i in range(2 * n + 1):
        dx = (X_pred[i] - x_pred).reshape(n, 1)
        P_pred += Wc[i] * (dx @ dx.T)
    P_pred = make_psd(P_pred + Qp)

    Z_pred = np.array([measurement_model_realistic(X_pred[i], I, current_time_s, interrogator_params, k_eff=None)[0] for i in range(2 * n + 1)])
    z_pred = np.sum(Wm[:, None] * Z_pred, axis=0)

    m = z_pred.shape[0]
    P_zz = np.zeros((m, m))
    P_xz = np.zeros((n, m))
    for i in range(2 * n + 1):
        dz = (Z_pred[i] - z_pred).reshape(m, 1)
        dx = (X_pred[i] - x_pred).reshape(n, 1)
        P_zz += Wc[i] * (dz @ dz.T)
        P_xz += Wc[i] * (dx @ dz.T)
    P_zz = make_psd(P_zz + R_measure)
    try:
        invPzz = np.linalg.inv(P_zz)
    except np.linalg.LinAlgError:
        invPzz = np.linalg.pinv(P_zz)

    K = P_xz @ invPzz
    x_upd = x_pred + K @ (z - z_pred)
    P_upd = P_pred - K @ P_zz @ K.T
    return np.nan_to_num(x_upd), make_psd(P_upd)

# ----------------------------- SECTION: FUSION -----------------------------
def fused_estimate_updated(x_ekf, x_ukf, P_ekf, P_ukf, R_measure=None, method='A'):
    if R_measure is not None:
        meas_trace = np.trace(R_measure)
        scale = meas_trace * 1e3
    else:
        scale = 0.0
    P_ekf_adj = P_ekf + np.eye(P_ekf.shape[0]) * scale
    P_ukf_adj = P_ukf + np.eye(P_ukf.shape[0]) * scale

    if method == 'A':
        w_ekf = 1.0 / np.trace(P_ekf_adj)
        w_ukf = 1.0 / np.trace(P_ukf_adj)
    elif method == 'B':
        w_ekf = 1.0 / (np.trace(P_ekf_adj) + 1e-12)
        w_ukf = 1.0 / (np.trace(P_ukf_adj) + 1e-12)
    elif method == 'C':
        w_ekf = 1.0 if x_ekf[0] > 0.5 else 0.0
        w_ukf = 1.0 - w_ekf
    else:
        w_ekf = 1.0 / (np.trace(P_ekf_adj) + 1e-12)
        w_ukf = 1.0 / (np.trace(P_ukf_adj) + 1e-12)

    w_sum = w_ekf + w_ukf
    if w_sum == 0:
        w_ekf = w_ukf = 0.5
    else:
        w_ekf /= w_sum
        w_ukf /= w_sum

    return w_ekf * x_ekf + w_ukf * x_ukf

# ----------------------------- SECTION: UTIL - SAVE METADATA -----------------------------
def save_run_metadata(csv_path, C_core_eff, C_surf_eff, R_th_eff, interrogator_params_run, ageing_factor_run, alpha_T_run):
    meta = {
        "C_core_eff_J_per_K": float(C_core_eff),
        "C_surf_eff_J_per_K": float(C_surf_eff),
        "R_th_eff_K_per_W": float(R_th_eff),
        "ageing_factor": float(ageing_factor_run),
        "alpha_T_run": float(alpha_T_run),
        "interrogator": {
            "b_lambda_pm": interrogator_params_run["b_lambda_m"] / pm_to_m,
            "b0_pm": interrogator_params_run["b0_m"] / pm_to_m,
            "drift_rate_pm_per_hour": fbg_drift_pm_per_hour,
            "eta_bond_0": interrogator_params_run["eta_bond_0"],
            "c_TS": interrogator_params_run["c_TS"],           
            
        }
    }
    meta_path = csv_path + ".meta.json"
    with open(meta_path, "w") as f:
        json.dump(meta, f, indent=2)           

# ----------------------------- SECTION: MAIN MONTE CARLO LOOP -----------------------------
ESTIMATOR_MODE = 'COMBINED'   # 'EKF', 'UKF', 'COMBINED' etc.
COMBINED_METHOD = 'A'         # fusion method

wall_start_time = time.time()

for T_amb in T_amb_range:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for cycle_name in cycle_names:
        I_profile, speed_profile = generate_synthetic_profile(cycle_name, time_vec)

        for mc_idx in range(num_mc_runs):
            if seeded:
                np.random.seed(seed_value + mc_idx)

            C_core_eff, C_surf_eff, R_th_eff = sample_thermal_parameters()
            ageing_factor_run = sample_ageing_factor(max_age=0.25)
            interrogator_params_run = sample_fbg_interrogator_characteristics()
            alpha_T_run = float(np.random.uniform(alpha_R_min, alpha_R_max))
            k_eff_run = k_th_nominal * interrogator_params_run["eta_bond_0"]

            R_measure_run = build_R_measure(num_fbgs, fbg_adc_noise_pm, interrogator_params_run)

            # initial states: SoC = 100%, voltages zero, temps ambient
            x_true = np.array([1.0, 0.0, 0.0, float(T_amb), float(T_amb)])
            x_ekf = x_true.copy()
            x_ukf = x_true.copy()
            P_ekf = np.eye(5) * 0.01
            P_ukf = np.eye(5) * 0.01

            df_records = []
            sim_time_seconds = 0.0

            for k, I in enumerate(I_profile):
                x_true = state_transition_realistic(x_true, I, T_amb, C_core_eff, C_surf_eff, R_th_eff, ageing_factor_run)

                z_meas, z_true_fbgs = measurement_model_realistic(x_true, I, sim_time_seconds, interrogator_params_run, k_eff=k_eff_run)

                v_noise = np.random.normal(0.0, math.sqrt(R_voltage))
                z_meas[0] += v_noise

                current_time_seconds = sim_time_seconds

                if ESTIMATOR_MODE == 'EKF':
                    x_ekf, P_ekf = ekf_predict_update(x_ekf, P_ekf, z_meas, I, T_amb,
                                                      C_core_eff, C_surf_eff, R_th_eff, ageing_factor_run,
                                                      interrogator_params_run, current_time_seconds,
                                                      R_measure=R_measure_run, Qp=Q_process)
                    x_fused = x_ekf.copy(); P_fused = P_ekf.copy()
                elif ESTIMATOR_MODE == 'UKF':
                    x_ukf, P_ukf = ukf_predict_update(x_ukf, P_ukf, z_meas, I, T_amb,
                                                      C_core_eff, C_surf_eff, R_th_eff, ageing_factor_run,
                                                      R_measure=R_measure_run, Qp=Q_process,
                                                      interrogator_params=interrogator_params_run,
                                                      current_time_s=current_time_seconds)
                    x_fused = x_ukf.copy(); P_fused = P_ukf.copy()
                else:  # COMBINED/BOTH
                    x_ekf, P_ekf = ekf_predict_update(x_ekf, P_ekf, z_meas, I, T_amb,
                                                      C_core_eff, C_surf_eff, R_th_eff, ageing_factor_run,
                                                      interrogator_params_run, current_time_seconds,
                                                      R_measure=R_measure_run, Qp=Q_process)
                    x_ukf, P_ukf = ukf_predict_update(x_ukf, P_ukf, z_meas, I, T_amb,
                                                      C_core_eff, C_surf_eff, R_th_eff, ageing_factor_run,
                                                      R_measure=R_measure_run, Qp=Q_process,
                                                      interrogator_params=interrogator_params_run,
                                                      current_time_s=current_time_seconds)
                    x_fused = fused_estimate_updated(x_ekf, x_ukf, P_ekf, P_ukf, R_measure=R_measure_run, method=COMBINED_METHOD)
                    P_fused = (P_ekf + P_ukf) / 2.0

                df_records.append(np.concatenate((
                    [float(T_amb), float(k), float(I)],
                    x_true, x_ekf, x_ukf, x_fused
                )))

                sim_time_seconds += dt

            df_cycle = pd.DataFrame(df_records, columns=[
                "T_amb", "time", "I",
                "SoC_true", "V1_true", "V2_true", "T_core_true", "T_surf_true",
                "SoC_EKF", "V1_EKF", "V2_EKF", "T_core_EKF", "T_surf_EKF",
                "SoC_UKF", "V1_UKF", "V2_UKF", "T_core_UKF", "T_surf_UKF",
                "SoC_Fused", "V1_Fused", "V2_Fused", "T_core_Fused", "T_surf_Fused"
            ])

            for col in ["SoC_true", "SoC_EKF", "SoC_UKF", "SoC_Fused"]:
                if df_cycle[col].max() <= 1.05:
                    df_cycle[col] = df_cycle[col] * 100.0

            sheet_name = f"{cycle_name}_MC{mc_idx+1}"
            ws = wb.create_sheet(title=sheet_name[:31])
            for r in dataframe_to_rows(df_cycle, index=False, header=True):
                ws.append(r)

            csv_path = os.path.join(timeseries_folder, f"{cycle_name}_MC{mc_idx+1}_T{T_amb}.csv")
            df_cycle.to_csv(csv_path, index=False)

            # write metadata
            save_run_metadata(csv_path, C_core_eff, C_surf_eff, R_th_eff, interrogator_params_run, ageing_factor_run, alpha_T_run)

    excel_path = os.path.join(summary_folder, f"simulation_T{T_amb}.xlsx")
    wb.save(excel_path)

# ----------------- CONFIG -----------------
timeseries_folder = "timeseries"  # adjust if different
expected_state_length = 5         # SoC, V1, V2, Tc, Ts
MC_runs = 1                     # number of Monte Carlo runs
cycles = ["cycle1"]               # list your cycle names
states_labels = ["SoC", "V1", "V2", "Tc", "Ts"]

# ----------------- HELPER FUNCTIONS -----------------
def check_array(arr, name="Array"):
    arr = np.array(arr)
    issues = []
    if not np.issubdtype(arr.dtype, np.number):
        issues.append("non-numeric")
    if np.any(np.isnan(arr)):
        issues.append("NaN values")
    if arr.ndim > 1 and arr.shape[1] != expected_state_length:
        issues.append(f"unexpected shape {arr.shape}")
    if issues:
        print(f"⚠ {name} issues: {', '.join(issues)}")

def scan_df_records(df_records):
    print("Scanning in-memory df_records...")
    df_arr = np.array([np.ravel(r) for r in df_records])
    for i, rec in enumerate(df_arr):
        check_array(rec, f"df_records[{i}]")
    # Quick SoC summary
    soc_indices = [3, 8, 13, 18]  # x_true, x_ekf, x_ukf, x_fused
    print("\nSoC min/max per estimator:")
    for idx, label in zip(soc_indices, ["True", "EKF", "UKF", "Fused"]):
        soc = df_arr[:, idx]
        print(f"{label}: min={soc.min():.4f}, max={soc.max():.4f}")

def scan_timeseries_folder(folder):
    print("\nScanning timeseries folder...")
    if os.path.exists(folder):
        files = [f for f in os.listdir(folder) if f.endswith(".npy") or f.endswith(".csv")]
        if not files:
            print("⚠ No timeseries files found!")
        for f in files:
            file_path = os.path.join(folder, f)
            try:
                if f.endswith(".npy"):
                    data = np.load(file_path, allow_pickle=True)
                else:  # CSV
                    import pandas as pd
                    data = pd.read_csv(file_path).values
                check_array(data, f"File: {f}")
            except Exception as e:
                print(f"⚠ Could not read {f}: {e}")
    else:
        print(f"⚠ Timeseries folder '{folder}' does not exist!")

# ----------------- RUN SCAN -----------------
if 'df_records' in globals():
    scan_df_records(df_records)
else:
    print("⚠ df_records not found in workspace.")

scan_timeseries_folder(timeseries_folder)

# ================================================================
# METRICS EVALUATION: ME, MAE, RMSE (NO PLOTTING)
# Single Excel workbook output
# ================================================================

# ----------------------------- PATHS -----------------------------
timeseries_folder = "simulation_outputs/timeseries"
output_excel = os.path.join("simulation_outputs", "summary",
                            "performance_metrics_ME_MAE_RMSE.xlsx")

# ----------------------------- METRIC FUNCTIONS -----------------------------
def ME(true, est):
    return float(np.mean(est - true))

def MAE(true, est):
    return float(np.mean(np.abs(est - true)))

def RMSE(true, est):
    return float(np.sqrt(np.mean((est - true) ** 2)))

# ----------------------------- EVALUATION -----------------------------
records = []

csv_files = sorted(glob.glob(os.path.join(timeseries_folder, "*.csv")))

# Define states and corresponding estimator columns
states = {
    "SoC": ("SoC_true", {
        "EKF": "SoC_EKF",
        "UKF": "SoC_UKF",
        "EKF-UKF": "SoC_Fused"
    }),
    "Core Temperature": ("T_core_true", {
        "EKF": "T_core_EKF",
        "UKF": "T_core_UKF",
        "EKF-UKF": "T_core_Fused"
    }),
    "Surface Temperature": ("T_surf_true", {
        "EKF": "T_surf_EKF",
        "UKF": "T_surf_UKF",
        "EKF-UKF": "T_surf_Fused"
    })
}

for csv_path in csv_files:
    df = pd.read_csv(csv_path)

    # Extract identifiers from filename
    fname = os.path.basename(csv_path)
    cycle = fname.split("_MC")[0]
    mc_id = int(fname.split("_MC")[1].split("_")[0])

    for state_name, (true_col, est_dict) in states.items():
        true_vals = df[true_col].values.astype(float)

        for estimator, est_col in est_dict.items():
            est_vals = df[est_col].values.astype(float)

            # Convert SoC from percent to fraction for evaluation
            if state_name == "SoC":
                true_vals_frac = true_vals / 100.0
                est_vals_frac  = est_vals  / 100.0
            else:
                true_vals_frac = true_vals
                est_vals_frac  = est_vals

            records.append({
                "Cycle": cycle,
                "Monte Carlo": mc_id,
                "Estimator": estimator,
                "State": state_name,
                "ME": ME(true_vals_frac, est_vals_frac),
                "MAE": MAE(true_vals_frac, est_vals_frac),
                "RMSE": RMSE(true_vals_frac, est_vals_frac)
            })

df_all = pd.DataFrame(records)

# ----------------------------- MONTE CARLO MEAN (PER CYCLE) -----------------------------
df_mc_mean = df_all.groupby(["Cycle", "Estimator", "State"], as_index=False).mean(numeric_only=True)

# ----------------------------- GLOBAL MEAN (ALL CYCLES & MCs) -----------------------------
df_global_mean = df_all.groupby(["Estimator", "State"], as_index=False).mean(numeric_only=True)

# ----------------------------- EXPORT TO EXCEL -----------------------------
os.makedirs(os.path.dirname(output_excel), exist_ok=True)
with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
    df_all.to_excel(writer, sheet_name="Per_MC_Per_Cycle", index=False)
    df_mc_mean.to_excel(writer, sheet_name="MC_Mean_Per_Cycle", index=False)
    df_global_mean.to_excel(writer, sheet_name="Global_Mean_All_Cycles", index=False)

print(f"\nMetrics successfully written to:\n{output_excel}")
# -----------------------------
def run_simulation(params, I_load_profile, T_amb):
    """
    Runs a single deterministic simulation using the full electro-thermal
    EKF–FBG–T2 model. Used for sensitivity analysis.
    """

    # -----------------------------
    # Thermal & ageing parameters
    # -----------------------------
    C_core_eff = params["C_th_core_nominal"]
    C_surf_eff = params["C_th_surf_nominal"]
    R_th_eff   = params["R_th_nominal"]
    ageing_factor = params.get("alpha_R_aging", 0.0)

    # -----------------------------
    # Interrogator parameters
    # (use canonical sampler)
    # -----------------------------
    interrogator_params = sample_fbg_interrogator_characteristics()

    # Override only parameters under sensitivity
    interrogator_params["eta_bond_0"] = params["eta_bond_nominal"]
    interrogator_params["c_TS"]       = params["c_TS_nominal"]

    # Effective thermal transfer
    k_eff = params["k_th_nominal"] * interrogator_params["eta_bond_0"]

    # -----------------------------
    # Initial states
    # -----------------------------
    x_true = np.array([1.0, 0.0, 0.0, float(T_amb), float(T_amb)])
    x_est  = x_true.copy()
    P_est  = np.eye(5) * 0.01

    # -----------------------------
    # Storage
    # -----------------------------
    SoC_true, SoC_est = [], []
    Tcore_true, Tcore_est = [], []
    Tsurf_true, Tsurf_est = [], []

    sim_time_seconds = 0.0
    dt_local = 1.0

    # -----------------------------
    # Time integration
    # -----------------------------
    for I in I_load_profile:

        # True propagation
        x_true = state_transition_realistic(
            x_true, I, T_amb,
            C_core_eff, C_surf_eff, R_th_eff,
            ageing_factor
        )

        # Measurements
        z_meas, _ = measurement_model_realistic(
            x_true, I, sim_time_seconds,
            interrogator_params,
            k_eff=k_eff
        )

        # EKF estimation
        x_est, P_est = ekf_predict_update(
            x_est, P_est, z_meas, I, T_amb,
            C_core_eff, C_surf_eff, R_th_eff,
            ageing_factor,
            interrogator_params,
            sim_time_seconds,
            R_measure=None,
            Qp=Q_process
        )

        # Record
        SoC_true.append(x_true[0] * 100.0)
        SoC_est.append(x_est[0] * 100.0)

        Tcore_true.append(x_true[3])
        Tcore_est.append(x_est[3])

        Tsurf_true.append(x_true[4])
        Tsurf_est.append(x_est[4])

        sim_time_seconds += dt_local

    return {
        "SoC_true": np.array(SoC_true),
        "SoC_est": np.array(SoC_est),
        "T_core_true": np.array(Tcore_true),
        "T_core_est": np.array(Tcore_est),
        "T_surf_true": np.array(Tsurf_true),
        "T_surf_est": np.array(Tsurf_est)
    }

# -----------------------------
# Error metrics
# -----------------------------
def compute_error_metrics(true, est):
    ME = np.mean(est - true)
    MAE = np.mean(np.abs(est - true))
    RMSE = np.sqrt(np.mean((est - true)**2))
    return ME, MAE, RMSE

# -----------------------------
# Full parameter table
# -----------------------------
sensitivity_params = {
    # FBG / sensor
    'lambda_B0': {'nominal': 1550e-9, 'range': [1550e-9*0.95, 1550e-9*1.05]},
    'rho_e': {'nominal': 0.6, 'range': [0.55, 0.65]},
    'alpha_fbg': {'nominal': 0.55e-6, 'range': [0.44e-6, 0.66e-6]},
    'xi': {'nominal': 6.7e-6, 'range': [5.36e-6, 8.04e-6]},
    'eta_bond_nominal': {'nominal': 0.8, 'range': [0.6, 1.0]},
    'tau_creep': {'nominal': 48*3600, 'range': [24*3600, 72*3600]},
    'fbg_adc_noise_pm': {'nominal': 3.0, 'range': [1.0, 5.0]},
    'c_TS_nominal': {'nominal': 0.15, 'range': [0.05, 0.2]},

    # Electrical
    'R0_ref_base': {'nominal': 0.015, 'range': [0.0135, 0.0165]},
    'R1_ref_base': {'nominal': 0.01, 'range': [0.009, 0.011]},
    'R2_ref_base': {'nominal': 0.02, 'range': [0.018, 0.022]},
    'C1_base': {'nominal': 2400.0, 'range': [2160, 2640]},
    'C2_base': {'nominal': 4800.0, 'range': [4080, 5520]},
    'alpha_R_base': {'nominal': 0.0085, 'range': [0.006, 0.018]},
    'alpha_R_soc': {'nominal': -0.15, 'range': [-0.2, -0.1]},
    'alpha_R_aging': {'nominal': 0.35, 'range': [0.25, 0.45]},

    # Thermal
    'C_th_core_nominal': {'nominal': 400.0, 'range': [340, 460]},
    'C_th_surf_nominal': {'nominal': 200.0, 'range': [170, 230]},
    'R_th_core_surf': {'nominal': 0.8, 'range': [0.64, 0.96]},
    'R_th_nominal': {'nominal': 3.0, 'range': [2.4, 3.6]},
    'k_th_nominal': {'nominal': 8e-7, 'range': [4e-7, 1.2e-6]},

    # Process / measurement noise
    'Q_process_scale': {'nominal': 1.0, 'range': [0.5, 1.5]},  # scale factor
    'R_voltage_scale': {'nominal': 3.0, 'range': [2.5, 3.5]},  

    # Model refinements
    'capacity_temp_coef': {'nominal': -2.2e-4, 'range': [-3.3e-4, -1.1e-4]},
    'ocv_temp_coef': {'nominal': -2.0e-4, 'range': [-3.0e-4, -1.0e-4]},
    'k_I_polar': {'nominal': 7e-3, 'range': [5e-3, 1e-2]},
    'low_soc_threshold': {'nominal': 0.15, 'range': [0.1, 0.2]},
    'low_soc_max_scale': {'nominal': 2.0, 'range': [1.5, 3.0]},
    'hyst_h0': {'nominal': 0.015, 'range': [0.0075, 0.0225]},
    'hyst_kI': {'nominal': 0.008, 'range': [0.004, 0.012]},
    'hyst_tau': {'nominal': 60.0, 'range': [30.0, 120.0]},

    # Environmental / operational
    'T_amb': {'nominal': 25.0, 'range': [-20, 60]},
    'I_load_scale': {'nominal': 1.0, 'range': [0.9, 1.1]}
}

# -----------------------------
# Load profile
# -----------------------------
time_steps = 1000
I_load_profile = np.ones(time_steps) * 50  # 50A constant load

# -----------------------------
# One-at-a-Time (OAT) Sensitivity
# -----------------------------
results_oat = {}
for param, info in sensitivity_params.items():
    values = np.linspace(info['range'][0], info['range'][1], 5)
    param_results = []
    for v in values:
        sim_params = {p: pinfo['nominal'] for p,pinfo in sensitivity_params.items()}
        sim_params[param] = v
        # Adjust load or ambient if parameter
        T_amb_val = sim_params.get('T_amb', 25.0)
        I_load_val = I_load_profile * sim_params.get('I_load_scale', 1.0)
        sim_out = run_simulation(sim_params, I_load_val, T_amb_val)
        ME_soc, MAE_soc, RMSE_soc = compute_error_metrics(sim_out['SoC_true'], sim_out['SoC_est'])
        ME_core, MAE_core, RMSE_core = compute_error_metrics(sim_out['T_core_true'], sim_out['T_core_est'])
        ME_surf, MAE_surf, RMSE_surf = compute_error_metrics(sim_out['T_surf_true'], sim_out['T_surf_est'])
        param_results.append({'value': v, 'ME_soc': ME_soc, 'MAE_soc': MAE_soc, 'RMSE_soc': RMSE_soc,
                              'ME_core': ME_core, 'MAE_core': MAE_core, 'RMSE_core': RMSE_core,
                              'ME_surf': ME_surf, 'MAE_surf': MAE_surf, 'RMSE_surf': RMSE_surf})
    results_oat[param] = param_results

# -----------------------------
# Tornado Plot Example (RMSE)
# -----------------------------
plt.figure(figsize=(10,12))
for i, (param, res_list) in enumerate(results_oat.items()):
    rmse_max = max(r['RMSE_soc'] for r in res_list)
    rmse_min = min(r['RMSE_soc'] for r in res_list)
    plt.barh(i, rmse_max - rmse_min, left=rmse_min, color='skyblue')
plt.yticks(range(len(results_oat)), list(results_oat.keys()))
plt.xlabel("SoC RMSE (%)")
plt.title("OAT Sensitivity Tornado Plot (SoC RMSE)")
plt.grid(True, axis='x')
plt.show()

# -----------------------------
# Monte Carlo Sensitivity
# -----------------------------
n_mc = 1
results_mc = []
for i in range(n_mc):
    sim_params = {p: np.random.uniform(info['range'][0], info['range'][1]) for p, info in sensitivity_params.items()}
    T_amb_val = sim_params.get('T_amb',25.0)
    I_load_val = I_load_profile * sim_params.get('I_load_scale', 1.0)
    sim_out = run_simulation(sim_params, I_load_val, T_amb_val)
    ME_soc, MAE_soc, RMSE_soc = compute_error_metrics(sim_out['SoC_true'], sim_out['SoC_est'])
    ME_core, MAE_core, RMSE_core = compute_error_metrics(sim_out['T_core_true'], sim_out['T_core_est'])
    ME_surf, MAE_surf, RMSE_surf = compute_error_metrics(sim_out['T_surf_true'], sim_out['T_surf_est'])
    results_mc.append({'RMSE_soc': RMSE_soc, 'RMSE_core': RMSE_core, 'RMSE_surf': RMSE_surf})

# Boxplots
plt.figure(figsize=(8,5))
rmse_soc = [r['RMSE_soc'] for r in results_mc]
rmse_core = [r['RMSE_core'] for r in results_mc]
rmse_surf = [r['RMSE_surf'] for r in results_mc]
plt.boxplot([rmse_soc, rmse_core, rmse_surf], labels=['SoC','Core Temp','Surface Temp'])
plt.ylabel("RMSE")
plt.title("Monte Carlo Sensitivity Analysis")
plt.grid(True)
plt.show()

# SINGLE-PLOT VISUALISATION: MC MEAN PER CYCLE (ME, MAE, RMSE)
# ================================================================

import matplotlib.pyplot as plt

# Define consistent colors per estimator
EST_COLORS = {"EKF": "tab:blue", "UKF": "tab:orange", "EKF-UKF": "tab:green"}

def plot_mc_mean_per_cycle(df_mc_mean, estimator_list, state_list):
    """
    Plots Monte Carlo mean error metrics (ME, MAE, RMSE) for all cycles,
    for multiple estimators and states.

    estimator_list: list of strings, e.g., ["EKF", "UKF", "EKF-UKF"]
    state_list: list of strings, e.g., ["SoC", "Core Temperature", "Surface Temperature"]
    """
    metrics = ["ME", "MAE", "RMSE"]

    for state_name in state_list:
        plt.figure(figsize=(12, 6))

        for estimator_name in estimator_list:
            df_sel = df_mc_mean[
                (df_mc_mean["Estimator"] == estimator_name) &
                (df_mc_mean["State"] == state_name)
            ]

            # Sort cycles to maintain order
            cycles = sorted(df_sel["Cycle"].unique())

            for cycle in cycles:
                row = df_sel[df_sel["Cycle"] == cycle].iloc[0]
                plt.plot(
                    metrics,
                    [row[m] for m in metrics],
                    marker="o",
                    linestyle="-",
                    color=EST_COLORS.get(estimator_name, None),
                    label=f"{estimator_name} – {cycle}"
                )

        plt.xlabel("Error Metric")
        plt.ylabel("Error Magnitude")
        plt.title(f"Monte Carlo Mean Error per Cycle – {state_name}")
        plt.grid(True)
        plt.legend(fontsize=9, loc="best")
        plt.tight_layout()
        plt.show()

# ----------------------------- USAGE -----------------------------
estimators = ["EKF", "UKF", "EKF-UKF"]
states = ["SoC", "Core Temperature", "Surface Temperature"]

plot_mc_mean_per_cycle(df_mc_mean, estimators, states)